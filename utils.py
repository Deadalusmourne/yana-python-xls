#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Time    : 2018/1/17 
# __author__: caoge
import os
from xlutils.filter import BaseReader, BaseFilter, BaseWriter, process
from openpyxl.cell import Cell
from openpyxl.utils import range_boundaries, get_column_letter
import re
from copy import copy
from xlutils.filter import process,XLRDReader,XLWTWriter
import xlrd, xlutils
from xlwt.Style import default_style
import XlsJinja
from XlsJinja import MultipleIterationError



class Reader(BaseReader):

    def __init__(self, sheet_index, filename):
        self.sheet_index = sheet_index
        self.filename = filename
        self.row_i = 0        # 读指针
        self.col_i = 0
        self.row_j = 0        # 写指针
        self.col_j = 0

    def get_filepaths(self):
        return [os.path.abspath(self.filename)]

    def get_workbooks(self):
        for path in self.get_filepaths():
            yield (
                xlrd.open_workbook(
                    path,
                    formatting_info=1,
                    on_demand=False,
                    ragged_rows=True),
                os.path.split(path)[1]
            )

    def __call__(self, filter):
        """
        Once instantiated, a reader will be called and have the first
        filter in the chain passed to its :meth:`__call__` method.
        The implementation of this method
        should call the appropriate methods on the filter based on the
        cells found in the :class:`~xlrd.Book` objects returned from the
        :meth:`get_workbooks` method.
        """
        filter.start()
        for workbook, filename in self.get_workbooks():
            # filter.workbook(workbook, filename)  # 把rd传递给后面所有的filter
            # for sheet_x in range(workbook.nsheets):
            #     sheet = workbook.sheet_by_index(sheet_x)
            #     filter.sheet(sheet, sheet.name)  # 把sheet也传递过去
            #     for row_x in xrange(sheet.nrows):
            #         filter.row(row_x, row_x)  # row index直传给了第一个
            #         for col_x in xrange(sheet.row_len(row_x)):  # sheet.row_len(row_x)
            #             filter.cell(row_x, col_x, row_x, col_x)
            #     if workbook.on_demand:
            #         workbook.unload_sheet(sheet_x)
            filter.workbook(workbook, filename)
            sheet = workbook.sheet_by_index(self.sheet_index)
            filter.sheet(sheet, sheet.name)
            total_row = sheet.nrows
            while True:
                if self.row_i >= total_row:
                    break
                total_col = sheet.row_len(self.row_i)
                req_row = filter.row(self.row_i, self.row_j)  # 返回值来调整指针
                self.col_i = 0                         # 重置列游标为0
                self.col_j = 0
                while True:
                    # print self.row_i, self.col_i, self.row_j, self.col_j, '|', total_col
                    if self.col_i >= total_col:
                        break
                    req_col = filter.cell(self.row_i, self.col_i, self.row_j, self.col_j)
                    if req_col:                  # 调整游标
                        self.row_i += req_col[1]
                        self.row_j += req_col[2]
                        self.col_i += req_col[3]
                        self.col_j += req_col[4]
                        if req_col[0]:
                            break
                    self.col_i += 1
                    self.col_j += 1
                self.row_i += 1
                self.row_j += 1
            if workbook.on_demand:
                workbook.unload_sheet(self.sheet_index)
        filter.finish()


class Filter(BaseFilter):             # 行处理的过滤器
    """
    cell 返回 1 1 1 1 1  break，四个游标 
    """
    pending_row = None
    def __init__(self, filename, xljianja):
        self.filename = filename
        self.xljianja = xljianja

    def workbook(self,rdbook,wtbook_name):
        self.next.workbook(rdbook,self.filename+'.new')

    def row(self,rdrowx,wtrowx):
        self.pending_row = (rdrowx,wtrowx)

    def cell(self, rdrowx, rdcolx, wtrowx, wtcolx, *args, **kwargs):
        print rdrowx, rdcolx, wtrowx, wtcolx
        cell_type = self.rdsheet.cell(rdrowx, rdcolx).ctype
        """
        xlrd.XL_CELL_EMPTY,
        xlrd.XL_CELL_TEXT,
        xlrd.XL_CELL_NUMBER,
        xlrd.XL_CELL_DATE,
        xlrd.XL_CELL_BOOLEAN,
        xlrd.XL_CELL_ERROR,
        xlrd.XL_CELL_BLANK
        """
        if cell_type == xlrd.XL_CELL_TEXT:
            cell_value = self.rdsheet.cell(rdrowx, rdcolx).value
            req_ass = XlsJinja.XlsJinja.assert_text(cell_value)
            if req_ass:
                if req_ass['error']:
                    raise MultipleIterationError(req_ass['error'])
                else:
                    control_type = req_ass['type']
                    control_data = req_ass.get('data')
                    if control_type == 'tr':
                        print 'got in tr loop', control_data
                        """
                        1 本行不写  break 掉
                        2 调整写游标的row 减1
                        3 给循环标志位为True
                        """
                        temp, temp_vb = control_data
                        loop_count = len(self.xljianja.render_vb.get(temp_vb))
                        self.xljianja.setbit(0, 1)          # 记录进入循环row状态
                        self.xljianja.setbit(2, loop_count) # 记录要循环几次
                        self.xljianja.setbit(4, loop_count) # 记录剩余
                        self.xljianja.tr_loop = rdrowx      # 用于计算循环体宽度
                        # self.next.cell(rdrowx, rdcolx, wtrowx, wtcolx, method='row_skip')
                        print self.xljianja.status
                        return 1, 0, 0, -2, 0               # 是否break 四个游标的位移
                    elif control_type == 'tc':
                        """  这里不对tc处理，下个filter处理tc
                        1 循环体内本列不写 在else限制
                        2 调整写游标col 减1
                        3 给循环标志位为True
                        """
                        temp, temp_vb = control_data
                        loop_count = len(self.xljianja.render.get(temp_vb))
                        self.xljianja.setbit(1, 1)
                        self.xljianja.setbit(3, loop_count)
                        self.xljianja.setbit(5, loop_count)
                        self.xljianja.tc_loop = rdcolx
                        # self.next.cell(rdrowx, rdcolx, wtrowx, wtcolx, method='col_skip')
                        return 0, 0, 0, 0, -1
                    elif control_type == 'set':
                        xl_dict = XlsJinja.XlsJinja.__dict__
                        if control_data[0] in xl_dict:
                            raise MultipleIterationError(
                                'please check up row%s col%s, valid variable param' % (rdrowx, rdcolx))
                        try:
                            setattr(self.xljianja, control_data[0], control_data[1])
                        except:
                            raise MultipleIterationError('please check up row%s col%s SET control'%(rdrowx, rdcolx))
                        self.next.cell(rdrowx, rdcolx, wtrowx, wtcolx, cell_value='', modify_value=True)
                    elif control_type == 'trendfor':
                        """
                        1 判断是否最后一次循环  是：调整row游标到for后面，跳过本行 否：
                        """
                        if self.xljianja.getbit(2) == self.xljianja.getbit(4):
                            loop_width = rdrowx - self.xljianja.tr_loop - 1
                            self.xljianja.setbit(6, abs(loop_width))   # 设置宽度 下次游标经过返回到for下面那行
                        if self.xljianja.getbit(4) > 1:                # 仍在循环中
                            loop_width = self.xljianja.getbit(6)
                            self.xljianja.setbit(4, int(self.xljianja.getbit(4))-1)
                            return 0, -int(loop_width), 0, 0, 0             # 调整游标
                        else:
                            return 1, 0, 0, -1, 0
                    elif control_type == 'tcendfor':
                        if self.xljianja.getbit(3) == self.xljianja.getbit(5):
                            loop_width = rdcolx - self.xljianja.tc_loop - 1
                            self.xljianja.setbit(7, abs(loop_width))
                        if self.xljianja.getbit(5) > 1:
                            loop_width = self.xljianja.getbit(7)
                            self.xljianja.setbit(5, self.xljianja.getbit(5) - 1)
                            return 0, 0, -loop_width, 0, 0
                        else:
                            return 1, 0, 0, 0, -1
                    elif control_type == 'variable':
                        print 'got into variable', cell_value
                        cell_value = self.xljianja.render_vb.get(control_data, '')
                        if hasattr(self.xljianja, control_data):
                            cell_value = getattr(self.xljianja, control_data)
                        self.next.cell(rdrowx, rdcolx, wtrowx, wtcolx, cell_value=cell_value, modify_value=True)
                    else:
                        self.next.cell(rdrowx, rdcolx, wtrowx, wtcolx)
            else:
                self.next.cell(rdrowx, rdcolx, wtrowx, wtcolx)
        else:
            self.next.cell(rdrowx, rdcolx, wtrowx, wtcolx)


    def sheet(self, rdsheet, wtsheet_name):
        self.rdsheet = rdsheet
        self.next.sheet(rdsheet, wtsheet_name)

    def set_rdsheet(self,rdsheet):
        print self.name, 'set_rdsheet', rdsheet
        self.next.sheet(rdsheet)


class Filter2(BaseFilter):      # 列处理过滤器
    pending_row = None
    def __init__(self, filename):
        self.total_col = 0
        self.filename = filename

    def workbook(self,rdbook,wtbook_name):
        self.next.workbook(rdbook,self.filename+'.new')

    def row(self,rdrowx,wtrowx):
        self.pending_row = (rdrowx,wtrowx)

    def cell(self,rdrowx,rdcolx,wtrowx,wtcolx, *args, **kwargs):
        # print 'self.total_col', self.total_col
        self.next.cell(rdrowx, rdcolx, wtrowx, wtcolx)


class Writer(BaseWriter):
    def get_stream(self,filename):
        return file(filename,'wb')

    def cell(self, rdrowx, rdcolx, wtrowx, wtcolx, *args, **kwargs):
        """
        This should be called for every cell in the sheet being processed.

        :param rdrowx: the index of the row to be read from in the current sheet. 
        :param rdcolx: the index of the column to be read from in the current sheet. 
        :param wtrowx: the index of the row to be written to in the current output sheet. 
        :param wtcolx: the index of the column to be written to in the current output sheet. 
        """
        cell = self.rdsheet.cell(rdrowx, rdcolx)

        # print args
        if kwargs.get('modify_value'):
            print 'kwargs',kwargs
            value_req = kwargs.get('cell_value')
            if isinstance(value_req, list):
                cell_value, cell_type = value_req
                setattr(cell, 'value', cell_value)
                setattr(cell, 'ctype', cell_type)
            elif isinstance(value_req,unicode):
                setattr(cell, 'value', value_req)
                setattr(cell, 'ctype', xlrd.XL_CELL_TEXT)
            elif isinstance(value_req, int):
                setattr(cell, 'value', value_req)
                setattr(cell, 'ctype', xlrd.XL_CELL_NUMBER)
            else:
                raise MultipleIterationError('%s data type error' % value_req)
        # setup column attributes if not already set
        if wtcolx not in self.wtcols and rdcolx in self.rdsheet.colinfo_map:
            rdcol = self.rdsheet.colinfo_map[rdcolx]
            wtcol = self.wtsheet.col(wtcolx)
            wtcol.width = rdcol.width
            wtcol.set_style(self.style_list[rdcol.xf_index])
            wtcol.hidden = rdcol.hidden
            wtcol.level = rdcol.outline_level
            wtcol.collapsed = rdcol.collapsed
            self.wtcols.add(wtcolx)
        # copy cell
        cty = cell.ctype
        if cty == xlrd.XL_CELL_EMPTY:
            return
        if cell.xf_index is not None:
            style = self.style_list[cell.xf_index]
        else:
            style = default_style
        rdcoords2d = (rdrowx, rdcolx)
        if rdcoords2d in self.merged_cell_top_left_map:
            # The cell is the governing cell of a group of
            # merged cells.
            rlo, rhi, clo, chi = self.merged_cell_top_left_map[rdcoords2d]
            assert (rlo, clo) == rdcoords2d
            self.wtsheet.write_merge(
                wtrowx, wtrowx + rhi - rlo - 1,
                wtcolx, wtcolx + chi - clo - 1,
                cell.value, style)
            return
        if rdcoords2d in self.merged_cell_already_set:
            # The cell is in a group of merged cells.
            # It has been handled by the write_merge() call above.
            # We avoid writing a record again because:
            # (1) It's a waste of CPU time and disk space.
            # (2) xlwt does not (as at 2007-01-12) ensure that only
            # the last record is written to the file.
            # (3) If you write a data record for a cell
            # followed by a blank record for the same cell,
            # Excel will display a blank but OOo Calc and
            # Gnumeric will display the data :-(
            return
        wtrow = self.wtsheet.row(wtrowx)
        if cty == xlrd.XL_CELL_TEXT:
            wtrow.set_cell_text(wtcolx, cell.value, style)
        elif cty == xlrd.XL_CELL_NUMBER or cty == xlrd.XL_CELL_DATE:
            wtrow.set_cell_number(wtcolx, cell.value, style)
        elif cty == xlrd.XL_CELL_BLANK:
            wtrow.set_cell_blank(wtcolx, style)
        elif cty == xlrd.XL_CELL_BOOLEAN:
            wtrow.set_cell_boolean(wtcolx, cell.value, style)
        elif cty == xlrd.XL_CELL_ERROR:
            wtrow.set_cell_error(wtcolx, cell.value, style)
        else:
            raise Exception(
                "Unknown xlrd cell type %r with value %r at (sheet=%r,rowx=%r,colx=%r)" \
                % (cty, cell.value, self.rdsheet.name, rdrowx, rdcolx)
            )

    # def set_rdsheet(self, rdsheet):
    #     print 'wawa', rdsheet


if __name__ == '__main__':
    xl = XlsJinja.XlsJinja()
    xl.render({'gaga': 111, 'vb': [{'a': 1, 'b': '2'}, {'a': '大范围', 'b': 'dfgre'}, {'a': 324, 'b': '的发个啥'}]})
    filename = 'test.xls'
    process(Reader(0, filename), Filter(filename, xl), Writer())
    os.rename(filename + ".new", 'new.xls')
    ## useage
    # xlsx insert

    # Worksheet.insert_rows = insert_rows
    # wb = openpyxl.load_workbook(filename='test.xlsx')
    # ws = wb.worksheets[0]
    # insert_rows(ws, 2, 5)
    # wb.save("testalpha.xlsx")

    # copy style

    # copy_style('test.xls.old', [1, 3], [20, 9], '', sheet=0)

    # insert xls


    # other
    # wb = xlrd.open_workbook('test.xls.old', on_demand=True)
    # wss = wb.
    # print wss
    # req = re.compile(r'(\\)')
    # req2 = req.search('asd\kyg')
    # print req2.groups()
